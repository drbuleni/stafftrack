"""Export utilities for PDF and Excel generation."""
from io import BytesIO
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def create_excel_report(title, headers, data, filename_prefix):
    """
    Create an Excel workbook with formatted data.

    Args:
        title: Report title
        headers: List of column headers
        data: List of rows (each row is a list of values)
        filename_prefix: Prefix for the filename

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet names max 31 chars

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
    date_cell.alignment = Alignment(horizontal="center")
    date_cell.font = Font(italic=True, size=10)

    # Headers row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_pdf_report(title, headers, data, filename_prefix, orientation='portrait'):
    """
    Create a PDF report with formatted data.

    Args:
        title: Report title
        headers: List of column headers
        data: List of rows (each row is a list of values)
        filename_prefix: Prefix for the filename
        orientation: 'portrait' or 'landscape'

    Returns:
        BytesIO buffer containing the PDF file
    """
    buffer = BytesIO()

    pagesize = landscape(A4) if orientation == 'landscape' else A4
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, topMargin=20*mm, bottomMargin=20*mm)

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#16a34a'),
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )

    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 10))

    # Table data
    table_data = [headers] + data

    # Calculate column widths
    page_width = pagesize[0] - 40*mm
    col_width = page_width / len(headers)

    table = Table(table_data, colWidths=[col_width] * len(headers))

    # Table styling
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a34a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def format_currency(amount):
    """Format amount as South African Rand."""
    if amount is None:
        return "R0.00"
    return f"R{float(amount):,.2f}"


def format_date(dt):
    """Format date for display."""
    if dt is None:
        return "-"
    if isinstance(dt, str):
        return dt
    return dt.strftime('%d %b %Y')


def format_datetime(dt):
    """Format datetime for display."""
    if dt is None:
        return "-"
    if isinstance(dt, str):
        return dt
    return dt.strftime('%d %b %Y %H:%M')
